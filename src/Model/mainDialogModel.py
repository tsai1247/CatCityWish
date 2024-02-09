import json, os, openpyxl
from time import sleep
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

from Common.UserInfo import UserInfo
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QColor


class Pools():
    poolNameList = ['常規召喚', '祈願召喚']
    others = '活動召喚'

    def getName(poolName: str) -> str:
        if poolName in Pools.poolNameList:
            return poolName
        else:
            return Pools.others

class MainDialogModel():
    def listToSummary(self):
        result = {}

        list = self.getList()
        for data in list:
            rarity = self.rarity[data['rare']]
            poolname = Pools.getName(data['poolName'])

            if poolname not in result:
                result[poolname] = {}
            if rarity not in result[poolname]:
                result[poolname][rarity] = 0
            
            result[poolname][rarity] += 1
            # result[poolname]['TotalCount'] += 1
        
        open('Summary.json', 'w', encoding='utf-8').write(json.dumps(result))
        return result

    def __init__(self) -> None:
        self.rarity = {
            1: 'R',
            2: 'SR',
            3: 'SSR',
            4: 'UR',
        }
        self.color = {
            'SSR': QColor(255, 255, 0, 255),
            'SR': QColor(124, 11, 255, 255),
            'R': QColor(0, 255, 241, 255) 
        }

        self.lastUpdatePrepend = '上次更新時間: '
        self.btn_Hint_Content = '操作說明'
        self.btn_UpdateData_Content = '更新資料'
        self.btn_ExportExcel_Content = '匯出Excel'

        self.userInfo = UserInfo()
        self.getData()

    def getData(self):
        summonSummary = self.listToSummary()

        for key in summonSummary.keys():
            summonSummary[key] = dict(sorted(summonSummary[key].items(), reverse=True))

        self.specialSummonTitle = '活動召喚'
        self.commonSummonTitle = '常規召喚'
        self.SSRSummonTitle = '祈願召喚'

                
        if summonSummary == {} or self.specialSummonTitle not in summonSummary:
            self.specialSummonData = []
        else:
            self.specialSummonData = list(map(
                lambda data: {
                    'name': f'{data[0]}: {data[1]}',
                    'value': data[1],
                    'color': self.color[data[0]],
                    'exploded': data[0] == 'SSR'
                },
                list(summonSummary[self.specialSummonTitle].items())
            ))

        if summonSummary == {} or self.commonSummonTitle not in summonSummary:
            self.commonSummonData = []
        else:
            self.commonSummonData = list(map(
                lambda data: {
                    'name': f'{data[0]}: {data[1]}',
                    'value': data[1],
                    'color': self.color[data[0]],
                    'exploded': data[0] == 'SSR'
                },
                list(summonSummary[self.commonSummonTitle].items())
            ))

        if summonSummary == {} or self.SSRSummonTitle not in summonSummary:
            self.SSRSummonData = []
        else:
            self.SSRSummonData = list(map(
                lambda data: {
                    'name': f'{data[0]}: {data[1]}',
                    'value': data[1],
                    'color': self.color[data[0]],
                    'exploded': data[0] == 'SSR'
                },
                list(summonSummary[self.SSRSummonTitle].items())
            ))

    def getList(self):
        load_dotenv()
        uid = os.getenv('playerId')
        if hasattr(self, 'userInfo') and hasattr(self.userInfo, 'playerId'):
            uid = self.userInfo.playerId
        if not uid:
            return []

        if not os.path.isfile(f'list{uid}.txt'):
            return []
        return json.loads(open(f'list{uid}.txt', 'r', encoding='utf-8').read())
    def setList(self, list):
        uid = self.userInfo.playerId
        open(f'list{uid}.txt', 'w', encoding='utf-8').write(json.dumps(list))


    def updateList(self) -> list:
        self.userInfo.login()


        # 取得舊資料
        oldList = self.getList()
        if len(oldList) > 0:
            lastCharacter = oldList[0]
        else:
            lastCharacter = None
        
        # 取得新資料
        pageSize = 10
        i = 0
        newList = []
        total = 0
        while i == 0 or pageSize * i < total:
            i += 1
            total, appendList = self.userInfo.callLog(i, pageSize)
            sleep(0.3)
            duplicated = False
            for character in appendList:
                if lastCharacter is not None and character == lastCharacter:
                    duplicated = True
                    break
                else:
                    newList.append(character)
            if duplicated: break

        # 合併新舊資料
        newList.extend(oldList)
        self.setList(newList)
        return newList

    def exportExcel(self, filename = None, list = None):
        if list is None:
            list = self.getList()

        # 開啟Excel
        print('\n整理中......')
        now = datetime.now().strftime('%Y%m%d-%H%M%S')
        wb = openpyxl.Workbook()

        commonSheet = wb.create_sheet("常規召喚")
        specialSheet =  wb.create_sheet("活動召喚")
        SSRSheet =  wb.create_sheet("祈願召喚")
        del wb["Sheet"]
            
        title = ['時間', '名稱', '稀有度', '總次數', '保底內次數', '備註']

        # set title
        for i in range(len(title)):
            commonSheet.cell(1, i+1).value = title[i]
            specialSheet.cell(1, i+1).value = title[i]
            SSRSheet.cell(1, i+1).value = title[i]

        commonCnt = 1
        commonGuaranteedCnt = 1
        commonSSRCnt = 0
        specialCnt = 1
        specialGuaranteedCnt = 1
        specialSSRCnt = 0
        SSRCnt = 1
        SSRGuaranteedCnt = 0

        list = list[::-1]
        
        for i in range(len(list)):
            card = list[i]
            if card['poolName'] == '常規召喚':
                values = [
                    card['createTime'], 
                    card['cardName'], 
                    self.rarity[card['rare']], 
                    commonCnt, 
                    commonGuaranteedCnt
                ]
                for j in range(len(values)):
                    commonSheet.cell(commonCnt + 1, j + 1).value = values[j]
                
                # fill
                if card['rare'] == 2:
                    for j in range(len(values)):
                        commonSheet.cell(commonCnt + 1, j + 1).font = Font(name='Arial', color='6F00D2')
                elif card['rare'] == 3:
                    for j in range(len(values)):
                        commonSheet.cell(commonCnt + 1, j + 1).font = Font(name='Arial', color='FFD306', bold=True)

                commonCnt += 1
                commonGuaranteedCnt += 1
                if card['rare'] > 2:
                    commonGuaranteedCnt = 0
                    commonSSRCnt += 1
            elif card['poolName'] == '祈願召喚':
                values = [
                    card['createTime'], 
                    card['cardName'], 
                    self.rarity[card['rare']], 
                    SSRCnt, 
                    '-'
                ]
                for j in range(len(values)):
                    SSRSheet.cell(SSRCnt + 1, j + 1).value = values[j]
                
                # fill
                if card['rare'] == 2:
                    for j in range(len(values)):
                        SSRSheet.cell(SSRCnt + 1, j + 1).font = Font(name='Arial', color='6F00D2')
                elif card['rare'] == 3:
                    for j in range(len(values)):
                        SSRSheet.cell(SSRCnt + 1, j + 1).font = Font(name='Arial', color='FFD306', bold=True)

                SSRCnt += 1
                    
            else:
                values = [
                    card['createTime'], 
                    card['cardName'], 
                    self.rarity[card['rare']], 
                    specialCnt, 
                    specialGuaranteedCnt
                ]
                for j in range(len(values)):
                    specialSheet.cell(specialCnt + 1, j + 1).value = values[j]
                
                # fill
                if card['rare'] == 2:
                    for j in range(len(values)):
                        specialSheet.cell(specialCnt + 1, j + 1).font = Font(name='Arial', color='6F00D2')
                elif card['rare'] == 3:
                    for j in range(len(values)):
                        specialSheet.cell(specialCnt + 1, j + 1).font = Font(name='Arial', color='FFD306', bold=True)

                specialCnt += 1
                specialGuaranteedCnt += 1
                if card['rare'] > 2:
                    specialGuaranteedCnt = 0
                    specialSSRCnt += 1

        # fill failed
        for i in range(6):
            commonSheet.column_dimensions[get_column_letter(i+1)].auto_size = True
            col = commonSheet.column_dimensions[get_column_letter(i+1)]
            col.fill = PatternFill(fill_type="solid", fgColor="8E8E8E")

        # total count 
        commonSheet.cell(1, 7).value = f'總抽數: {commonCnt}'
        commonSheet.cell(1, 8).value = f'SSR數量: {commonSSRCnt}'
        commonSheet.cell(1, 9).value = f'機率: {commonSSRCnt * 10000 // commonCnt / 100}%'
        specialSheet.cell(1, 7).value = f'總抽數: {specialCnt}'
        specialSheet.cell(1, 8).value = f'SSR數量: {specialSSRCnt}'
        specialSheet.cell(1, 9).value = f'機率: {specialSSRCnt * 10000 // specialCnt / 100}%'
        
        if filename is None:
            filename = f'貓之城召喚紀錄_{now}.xlsx'
        wb.save(filename)
