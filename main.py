import glob
import base64
import json
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from time import sleep
from dotenv import load_dotenv
from os import getenv

load_dotenv()

# https://passport-user-center-pc.fundollgame.com/#/queryCenter

class UserInfo:
    def __init__(self, catcityPath='C:\Program Files\貓之城') -> None:
        userInfoPath = f'{catcityPath}\GameClient\CatFantasy\CatFantasy_Data'
        userInfoList = glob.glob(f'{userInfoPath}\*.txt')

        if len(userInfoList) == 0:
            raise RuntimeError("Path not found")

        userInfo_encoded = open(userInfoList[0], 'r', encoding='utf-8').read()
        userInfo = json.loads(base64.b64decode(userInfo_encoded))

        self.clientAppId = getenv('clientAppId')
        self.playerId = getenv('playerId')
        self.username = userInfo['username']
        self.accesstoken = userInfo['accesstoken']
        self.sign = userInfo['sign']
        self.openId = str(userInfo['openId'])

    def login(self, sessionId = None):
        if sessionId is not None:
            self.sessionId = sessionId
            self.expireAt = datetime.timestamp(datetime.now()) + 1800
            return
        
        URL = 'https://cat-tw.fungoglobal.com/webLogQuery/autoLogin'
        payload = {
            "clientAppId": self.clientAppId,
            "accessToken": self.accesstoken
        }

        res = requests.post(URL, json = payload)
        response = json.loads(res.text)
        self.sessionId = response["data"]["sessionId"]
        self.expireAt = response["data"]["expireAt"]

    def callLog(self, page = 1, pageSize = 10):
        print(f'\r正在抓取第{page}頁', end='')
        URL = 'https://cat-tw.fungoglobal.com/webLogQuery/callLog'
        payload = {
            "sessionId": self.sessionId,
            "playerId": self.playerId,
            "lang": "zh-tw",
            "pageSize": pageSize,
            "page": page
        }

        res = requests.post(URL, json = payload)
        response = json.loads(res.text)
        return response['data']['total'], response['data']['list']

rarity = {
    1: 'R',
    2: 'SR',
    3: 'SSR',
    4: 'UR',
}

if __name__ == '__main__':
    userInfo = UserInfo()
    userInfo.login()

    oldList = json.loads(open('list.txt', 'r', encoding='utf-8').read())
    lastCharacter = oldList[0]
    
    pageSize = 10
    i = 0
    newList = []
    total = 0
    while i == 0 or pageSize * i < total:
        i += 1
        total, appendList = userInfo.callLog(i, pageSize)
        sleep(0.3)
        duplicated = False
        for character in appendList:
            if character == lastCharacter:
                duplicated = True
                break
            else:
                newList.append(character)
        if duplicated: break

    newList.extend(oldList)
    open('list.txt', 'w', encoding='utf-8').write(json.dumps(newList))
    list = newList

    print('\n整理中......')

    now = datetime.now().strftime('%Y%m%d-%H%M%S')
    wb = openpyxl.Workbook()

    commonSheet = wb.create_sheet("常規召喚")
    specialSheet =  wb.create_sheet("活動召喚")
    SSRSheet =  wb.create_sheet("祈願召喚")
    del wb["Sheet"]
        
    title = ['時間', '名稱', '稀有度', '總次數', '保底內次數', '備註']

    # set title
    for i in range(len(title)):
        commonSheet.cell(1, i+1).value = title[i]
        specialSheet.cell(1, i+1).value = title[i]
        SSRSheet.cell(1, i+1).value = title[i]

    commonCnt = 1
    commonGuaranteedCnt = 1
    commonSSRCnt = 0
    specialCnt = 1
    specialGuaranteedCnt = 1
    specialSSRCnt = 0
    SSRCnt = 1
    SSRGuaranteedCnt = 0

    list = list[::-1]
    
    for i in range(len(list)):
        card = list[i]
        if card['poolName'] == '常規召喚':
            values = [
                card['createTime'], 
                card['cardName'], 
                rarity[card['rare']], 
                commonCnt, 
                commonGuaranteedCnt
            ]
            for j in range(len(values)):
                commonSheet.cell(commonCnt + 1, j + 1).value = values[j]
            
            # fill
            if card['rare'] == 2:
                for j in range(len(values)):
                    commonSheet.cell(commonCnt + 1, j + 1).font = Font(name='Arial', color='6F00D2')
            elif card['rare'] == 3:
                for j in range(len(values)):
                    commonSheet.cell(commonCnt + 1, j + 1).font = Font(name='Arial', color='FFD306', bold=True)

            commonCnt += 1
            commonGuaranteedCnt += 1
            if card['rare'] > 2:
                commonGuaranteedCnt = 0
                commonSSRCnt += 1
        elif card['poolName'] == '祈願召喚':
            values = [
                card['createTime'], 
                card['cardName'], 
                rarity[card['rare']], 
                SSRCnt, 
                '-'
            ]
            for j in range(len(values)):
                SSRSheet.cell(SSRCnt + 1, j + 1).value = values[j]
            
            # fill
            if card['rare'] == 2:
                for j in range(len(values)):
                    SSRSheet.cell(SSRCnt + 1, j + 1).font = Font(name='Arial', color='6F00D2')
            elif card['rare'] == 3:
                for j in range(len(values)):
                    SSRSheet.cell(SSRCnt + 1, j + 1).font = Font(name='Arial', color='FFD306', bold=True)

            SSRCnt += 1
                
        else:
            values = [
                card['createTime'], 
                card['cardName'], 
                rarity[card['rare']], 
                specialCnt, 
                specialGuaranteedCnt
            ]
            for j in range(len(values)):
                specialSheet.cell(specialCnt + 1, j + 1).value = values[j]
            
            # fill
            if card['rare'] == 2:
                for j in range(len(values)):
                    specialSheet.cell(specialCnt + 1, j + 1).font = Font(name='Arial', color='6F00D2')
            elif card['rare'] == 3:
                for j in range(len(values)):
                    specialSheet.cell(specialCnt + 1, j + 1).font = Font(name='Arial', color='FFD306', bold=True)

            specialCnt += 1
            specialGuaranteedCnt += 1
            if card['rare'] > 2:
                specialGuaranteedCnt = 0
                specialSSRCnt += 1

    # fill failed
    for i in range(6):
        commonSheet.column_dimensions[get_column_letter(i+1)].auto_size = True
        col = commonSheet.column_dimensions[get_column_letter(i+1)]
        col.fill = PatternFill(fill_type="solid", fgColor="8E8E8E")

    # total count 
    commonSheet.cell(1, 7).value = f'總抽數: {commonCnt}'
    commonSheet.cell(1, 8).value = f'SSR數量: {commonSSRCnt}'
    commonSheet.cell(1, 9).value = f'機率: {commonSSRCnt * 10000 // commonCnt / 100}%'
    specialSheet.cell(1, 7).value = f'總抽數: {specialCnt}'
    specialSheet.cell(1, 8).value = f'SSR數量: {specialSSRCnt}'
    specialSheet.cell(1, 9).value = f'機率: {specialSSRCnt * 10000 // specialCnt / 100}%'
    


    wb.save(f'貓之城召喚紀錄_{now}.xlsx')
